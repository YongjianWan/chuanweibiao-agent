"""把 S4 报告（report.json）导成 Excel。

**本次交付的重点不是分数本身，是「这一分从哪来」**（README §1：P1「判分有据可查」是硬要求，
P2「分数与专家一致」本次不追求）。所以表的排布刻意把依据放在主位：

| 工作表 | 内容 |
|---|---|
| 判分与依据 | 228 行，一行一个判分。**得分紧挨着判分理由和证据出处**，出处精确到「PDF 文件 · 第几页 · 章节路径」。这是本文件的主体 |
| 评分汇总 | 19 项 × 12 家矩阵 + 合计。给要看总分的人，**不是重点** |
| 未评定 | 系统没能判出来的项，单列（README §3.6：不得记 0 分、不混进分数统计） |
| 建议复核 | 证据降级项 + 无区分度项，附原因 |
| 性能与算力 | 耗时、调用数、token（标注估算）、GPU、算力归属 |

**证据出处怎么来的**：评审结果里的 `cite` 是证据包 `picked` 数组的下标（模型只选编号、
不生成原文，见 README §3.5 的防幻觉设计）。本脚本读 `evidence/<bidder>/located.json`
把下标还原成 `文件名 · p.页码 · 章节 > 路径`。**拿不到 evidence 目录时退化为只写编号**，
那种情况下「依据」这一列对甲方是没用的，应当补上 `--evidence` 参数重新导。

**三条渲染约定来自 README §4，不要改：**

1. **`0` 分与「未评定」必须视觉可区分**：0 分写 `0`，未评定写 `—`（全角破折号）。
   两者含义完全不同——0 分是投标人没写，未评定是系统没判出来。
2. **未评定不计入合计。** 合计行只加已判分的项。
3. **token 必须标注是本地估算**，端点不返回 usage（README §8 的 B2）。

用法：

    python scripts/export_xlsx.py data/runs/<run_id>/report/report.json
    python scripts/export_xlsx.py data/out/report/report.json -o 评审结果.xlsx

不传 `-o` 时输出到 report.json 同目录下的 `评审报告.xlsx`。
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 未评定的显示符号。与 0 分区分开，见模块 docstring 的约定 1。
UNRATED_MARK = "—"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, size=9, color="666666")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def short_name(bidder: str) -> str:
    """表头用短名：去掉尾部的投标编号数字，太长的列头会把表撑爆。"""
    trimmed = bidder.rstrip("0123456789")
    return trimmed or bidder


def load_evidence(evidence_dir: Path | None) -> dict:
    """读 evidence/<bidder>/located.json，建 (bidder, item_id) → picked 列表 的索引。

    评审结果里的 `cite` 是 picked 的下标——模型只选编号、不生成原文（README §3.5），
    所以要还原出处必须回到证据包。拿不到目录就返回空索引，出处列退化为编号。"""
    index: dict = {}
    if evidence_dir is None or not evidence_dir.exists():
        return index
    for located in sorted(evidence_dir.glob("*/located.json")):
        try:
            packages = json.loads(located.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        for pack in packages:
            key = (pack.get("bidder") or located.parent.name, pack.get("item_id"))
            index[key] = pack.get("picked") or []
    return index


def cite_sources(row, evidence_index) -> str:
    """把 cite 编号还原成人能读的出处，一行一条。

    形如：`设计任务书优化….pdf · p.15 · 第二章 施工组织设计 > 一、总体概况 > 1、达到…`
    甲方拿这一列就能翻回投标文件原文核对，这是 README §1 的 P1「判分有据可查」的落点。"""
    cites = row.get("cite") or []
    if not cites:
        return ""
    picked = evidence_index.get((row.get("bidder"), row.get("item_id")))
    if not picked:
        return "证据编号 " + "、".join(str(c) for c in cites) + "（未提供 evidence 目录，无法还原出处）"
    lines = []
    for idx in cites:
        if not isinstance(idx, int) or idx >= len(picked):
            lines.append(f"[{idx}] 编号越界")
            continue
        item = picked[idx]
        path = " > ".join(item.get("path") or [])
        page = item.get("page")
        lines.append(f"[{idx}] {item.get('file', '')} · "
                     f"{'p.' + str(page) if page else '页码未采集'} · {path}")
    return "\n".join(lines)


def sheet_evidence(wb, report, evidence_index) -> None:
    """主表：一行一个判分，得分紧挨着理由和出处。

    列序刻意如此——本次交付的重点是「这一分从哪来」，不是分数本身（README §1）。
    分数放在理由左边一列，读的人视线自然从「多少分」滑到「为什么」再到「哪来的」。"""
    ws = wb.create_sheet("判分与依据")
    ws.cell(row=1, column=1,
            value=f"{report.get('project', '')} · 判分与依据（共 {len(report['details'])} 项）").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="本次交付的重点是每个判分可追溯：得分右边是判分理由，再右边是该理由所依据的"
                  "投标文件出处（PDF · 页码 · 章节路径），可直接翻回原文核对。"
                  f"“{UNRATED_MARK}”表示未评定，与 0 分含义不同。").font = NOTE_FONT
    ws.append([])

    header = ["投标人", "评分项", "满分", "得分", "档位",
              "判分理由", "依据出处（投标文件 · 页码 · 章节）",
              "编号", "0 分原因", "置信度", "调用次数"]
    ws.append(header)
    head_row = ws.max_row
    style_header(ws, head_row, len(header))

    item_meta = {r["item_id"]: r for r in report["matrix"]}
    miss_label = {"no_file": "缺文件（该家确实没有这个文件）",
                  "not_found": "检索未命中（文件在，未定位到内容）——把「写了」判成「没写」的风险位"}

    for row in report["details"]:
        meta = item_meta.get(row.get("item_id"), {})
        rated = row.get("status") == "rated"
        ws.append([
            row.get("bidder", ""),
            meta.get("name", "") or row.get("item_id", ""),
            meta.get("max_score"),
            row.get("score") if rated else UNRATED_MARK,
            row.get("tier") or UNRATED_MARK,
            (row.get("reason") or "") if rated else
            f"未评定：{row.get('last_error') or '重试耗尽'}",
            cite_sources(row, evidence_index),
            "、".join(str(c) for c in (row.get("cite") or [])),
            miss_label.get(row.get("miss_reason"), ""),
            row.get("confidence"),
            row.get("attempts"),
        ])
        cells = [ws.cell(row=ws.max_row, column=c) for c in range(1, len(header) + 1)]
        for cell in cells:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top")
        cells[5].alignment = Alignment(wrap_text=True, vertical="top")
        cells[6].alignment = Alignment(wrap_text=True, vertical="top")
        if not rated:
            for cell in cells:
                cell.fill = WARN_FILL

    set_widths(ws, [24, 28, 6, 7, 7, 68, 62, 12, 30, 8, 9])
    ws.freeze_panes = ws.cell(row=head_row + 1, column=3)
    ws.auto_filter.ref = f"A{head_row}:{get_column_letter(len(header))}{ws.max_row}"


def sheet_summary(wb, report) -> None:
    ws = wb.create_sheet("评分汇总")
    bidders = report["bidders"]
    ws.cell(row=1, column=1, value=f"{report.get('project', '')} · 技术标辅助评审 评分汇总").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"生成时间 {report.get('generated_at', '')}　"
                  f"｜　{len(bidders)} 家投标人 × {len(report['matrix'])} 个评分项"
                  f"　｜　“{UNRATED_MARK}”表示未评定（系统未判出，不同于 0 分）").font = NOTE_FONT

    header = ["序号", "评分项", "满分"] + [short_name(b) for b in bidders]
    ws.append([])
    ws.append(header)
    head_row = ws.max_row
    style_header(ws, head_row, len(header))

    for idx, row in enumerate(report["matrix"], start=1):
        scores = row.get("scores") or {}
        line = [idx, row.get("name", ""), row.get("max_score")]
        for bidder in bidders:
            value = scores.get(bidder)
            line.append(UNRATED_MARK if value is None else value)
        ws.append(line)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = BORDER
            if col >= 3:
                cell.alignment = Alignment(horizontal="center")

    totals = report.get("totals") or {}
    total_line = ["", "合计（未评定项不计入）",
                  sum(r.get("max_score") or 0 for r in report["matrix"])]
    for bidder in bidders:
        total_line.append((totals.get(bidder) or {}).get("score"))
    ws.append(total_line)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True)
        cell.border = BORDER
        if col >= 3:
            cell.alignment = Alignment(horizontal="center")

    unrated_line = ["", "其中未评定项数", ""]
    for bidder in bidders:
        unrated_line.append((totals.get(bidder) or {}).get("unrated", 0))
    ws.append(unrated_line)
    for col in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=col).border = BORDER

    set_widths(ws, [6, 34, 7] + [13] * len(bidders))
    ws.freeze_panes = ws.cell(row=head_row + 1, column=4)


def sheet_unrated(wb, report) -> None:
    ws = wb.create_sheet("未评定")
    ws.cell(row=1, column=1,
            value="未评定 ≠ 0 分：0 分表示投标人没写，未评定表示系统没判出来（重试耗尽）。"
                  "两者不合并统计，见 README §3.6。").font = NOTE_FONT
    ws.append([])
    header = ["投标人", "评分项编号", "调用次数", "最后一次错误"]
    ws.append(header)
    style_header(ws, ws.max_row, len(header))

    rows = report.get("unrated") or []
    for row in rows:
        ws.append([row.get("bidder", ""), row.get("item_id", ""),
                   row.get("attempts"), row.get("last_error", "")])
        for col in range(1, len(header) + 1):
            ws.cell(row=ws.max_row, column=col).border = BORDER
    if not rows:
        ws.append(["（本次无未评定项）", "", "", ""])

    set_widths(ws, [26, 12, 10, 70])


def sheet_flags(wb, report) -> None:
    ws = wb.create_sheet("建议复核")
    ws.cell(row=1, column=1,
            value="系统标出、交人复核，不替人改分。两类：置信度偏低；某评分项各家判成同一档、区分不出来。").font = NOTE_FONT
    ws.append([])
    header = ["类型", "投标人", "评分项编号", "置信度 / 档位分布", "原因"]
    ws.append(header)
    style_header(ws, ws.max_row, len(header))

    for row in report.get("review_flags") or []:
        ws.append(["低置信", row.get("bidder", ""), row.get("item_id", ""),
                   row.get("confidence"), row.get("why", "")])
    for row in report.get("audit") or []:
        dist = row.get("tier_dist") or {}
        ws.append(["无区分度", "（全部投标人）", row.get("item_id", ""),
                   "　".join(f"{k} {v}" for k, v in dist.items()),
                   row.get("detail", "")])
    for r in range(4, ws.max_row + 1):
        for col in range(1, len(header) + 1):
            ws.cell(row=r, column=col).border = BORDER

    set_widths(ws, [12, 26, 12, 26, 52])
    ws.auto_filter.ref = f"A3:{get_column_letter(len(header))}{ws.max_row}"


def sheet_perf(wb, report) -> None:
    ws = wb.create_sheet("性能与算力")
    perf = report.get("perf") or {}
    notes = report.get("compute_notes") or {}

    def put(label, value, note=""):
        ws.append([label, value, note])
        for col in range(1, 4):
            ws.cell(row=ws.max_row, column=col).border = BORDER
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.cell(row=1, column=1, value="性能数据与算力说明").font = TITLE_FONT
    ws.append([])
    ws.append(["项目", "值", "说明"])
    style_header(ws, ws.max_row, 3)

    wall = perf.get("wall_clock_sec")
    put("总耗时", f"{wall} 秒（约 {round(wall / 60, 1)} 分钟）" if wall else "未采集",
        perf.get("wall_clock_note", ""))
    put("并发路数", perf.get("concurrency"), "逐项评审的并发数")
    put("评审项次数", perf.get("calls"), "覆盖度校验位 = 投标人数 × 评分项数")
    put("重试次数", perf.get("retries"), "调用失败后的额外重试，不是失败项数")
    put("输入 tokens", perf.get("in_tokens"),
        "**本地估算**：端点不返回 token 计数，按中文约 1.5 字/token 估算")
    put("输出 tokens", perf.get("out_tokens"), "同上，为估算值")
    put("GPU / 显存", f"{perf.get('gpu')} / {perf.get('vram_peak_gb') or '未采集'}",
        perf.get("gpu_note", ""))
    ws.append([])
    put("算力归属", notes.get("owner", "未采集"), "")
    put("硬件规格", notes.get("spec", "未采集"), "")
    put("模型版本", notes.get("model", "未采集"), "")

    method = notes.get("method") or []
    if method:
        ws.append([])
        ws.append(["技术做法", "", ""])
        style_header(ws, ws.max_row, 3)
        for line in method:
            ws.append(["", line, ""])
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)

    set_widths(ws, [18, 46, 62])


def build(report: dict, evidence_index: dict | None = None) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    # 顺序即优先级：依据在前、分数在后。见模块 docstring。
    sheet_evidence(wb, report, evidence_index or {})
    sheet_summary(wb, report)
    sheet_unrated(wb, report)
    sheet_flags(wb, report)
    sheet_perf(wb, report)
    return wb


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="把 S4 报告导成 Excel")
    parser.add_argument("report", type=Path, help="report.json 路径")
    parser.add_argument("-o", "--output", type=Path, default=None,
                        help="输出 xlsx 路径，默认与 report.json 同目录")
    parser.add_argument("--evidence", type=Path, default=None,
                        help="evidence 目录，用于把 cite 编号还原成投标文件出处。"
                             "不传时自动找 <report.json 的上两级>/evidence")
    args = parser.parse_args(argv)

    if not args.report.exists():
        print(f"找不到 {args.report}")
        return 1

    report = json.loads(args.report.read_text(encoding="utf-8"))

    # 默认按约定布局找 evidence：report.json 在 <run>/report/ 下，证据在 <run>/evidence/
    evidence_dir = args.evidence or args.report.parent.parent / "evidence"
    evidence_index = load_evidence(evidence_dir)

    output = args.output or args.report.parent / "评审报告.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    build(report, evidence_index).save(output)

    print(f"已生成 {output}")
    print(f"  {len(report.get('bidders', []))} 家 × {len(report.get('matrix', []))} 项"
          f"　未评定 {len(report.get('unrated') or [])} 项"
          f"　建议复核 {len(report.get('review_flags') or []) + len(report.get('audit') or [])} 项")
    if evidence_index:
        print(f"  证据出处已还原（读到 {len(evidence_index)} 个证据包）")
    else:
        print(f"  ⚠ 未读到证据包（找的是 {evidence_dir}），"
              "「依据出处」列只会写编号，甲方看不懂——请用 --evidence 指定目录重导")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
